import random
import string
import re
from typing import Tuple, Optional, List
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from aiogram import Bot
from aiogram.exceptions import TelegramBadRequest
import logging


async def send_test_content(target, content: str, parse_mode: str = None):
    """Send stored test content correctly.

    `target` is a `Message`-like object with `answer`, `answer_photo`, `answer_document` methods.
    `content` format: "photo:<file_id>[|caption:...]", or "pdf:<file_id>[|caption:...]", or plain text.
    """
    if not content:
        await target.answer("(No content)")
        return

    try:
        if content.startswith("photo:"):
            rest = content.split(":", 1)[1]
            file_id = rest.split("|", 1)[0]
            caption = None
            if "|caption:" in rest:
                caption = rest.split("|caption:", 1)[1]
            await target.answer_photo(photo=file_id, caption=caption or "", parse_mode=parse_mode)
            return

        if content.startswith("pdf:"):
            rest = content.split(":", 1)[1]
            file_id = rest.split("|", 1)[0]
            caption = None
            if "|caption:" in rest:
                caption = rest.split("|caption:", 1)[1]
            await target.answer_document(document=file_id, caption=caption or "", parse_mode=parse_mode)
            return

        # fallback to plain text
        await target.answer(content, parse_mode=parse_mode)
    except Exception as e:
        logging.exception("Failed to send test content: %s", e)
        # fallback to sending as text
        await target.answer(content)


async def check_user_subscription(bot: Bot, user_id: int, db) -> Tuple[bool, list]:
    """
    Returns: (is_subscribed, unsubscribed_channels)
    """
    channels = await db.get_all_channels()
    unsubscribed = []
    inaccessible = []

    for channel in channels:
        channel_id = None
        username = None
        try:
            if isinstance(channel, dict):
                channel_id = channel.get('channel_id')
                username = channel.get('username')
            else:
                channel_id = getattr(channel, 'channel_id', None)
                username = getattr(channel, 'username', None)

            if channel_id is None:
                logging.warning("Channel without id in DB: %s", channel)
                continue

            try:
                member = await bot.get_chat_member(channel_id, user_id)
                if member.status not in ['member', 'administrator', 'creator']:
                    unsubscribed.append(channel)
            except TelegramBadRequest as e:
                msg = str(e).lower()
                logging.warning("TelegramBadRequest for channel %s: %s", channel_id, msg)
                if 'member list is inaccessible' in msg or 'chat not found' in msg or 'have no rights' in msg:
                    inaccessible.append(channel)
                else:
                    unsubscribed.append(channel)
        except Exception as e:
            logging.exception("Error checking subscription for channel %s: %s", channel, e)

    all_blockers = unsubscribed + inaccessible
    return len(all_blockers) == 0, all_blockers



def generate_test_code() -> str:
    return ''.join(random.choices(string.digits, k=3))


async def get_unique_test_code(db) -> str:
    while True:
        code = generate_test_code()
        if not await db.code_exists(code):
            return code


def validate_answer_key(answer_key: str, question_count: int) -> Tuple[bool, Optional[str]]:
    """
    Format: 1a2b3c4d...
    Returns: (is_valid, error_message)
    """
    if not answer_key:
        return False, "Javob kaliti bo'sh bo'lmasligi kerak"
    
    if not re.match(r'^(\d+[a-d])+$', answer_key.lower()):
        return False, "Format xato! Masalan: 1a2b3c4a"
    
    pattern = r'(\d+)([a-d])'
    matches = re.findall(pattern, answer_key.lower())
    
    if len(matches) != question_count:
        return False, f"Savollar soni {question_count} ta bo'lishi kerak, lekin {len(matches)} ta topildi"
    
    for i, (num, _) in enumerate(matches, 1):
        if int(num) != i:
            return False, f"Savollar ketma-ket bo'lishi kerak (1, 2, 3...). {num}-savol xato"
    
    return True, None


def validate_user_answer(user_answer: str, question_count: int) -> Tuple[bool, Optional[str]]:

    if not user_answer:
        return False, "Javob bo'sh bo'lmasligi kerak"
    
    if not re.match(r'^(\d+[a-d])+$', user_answer.lower()):
        return False, "Format xato! Masalan: 1a2b3c4a"
    
    pattern = r'(\d+)([a-d])'
    matches = re.findall(pattern, user_answer.lower())
    
    if len(matches) != question_count:
        return False, f"Barcha {question_count} ta savolga javob bering"
    
    return True, None


def check_answers(user_answer: str, correct_answer: str) -> Tuple[int, int]:
    """
    Returns: (correct_count, wrong_count)
    """
    user_pattern = r'(\d+)([a-d])'
    user_matches = re.findall(user_pattern, user_answer.lower())
    correct_matches = re.findall(user_pattern, correct_answer.lower())
    
    user_dict = {int(num): letter for num, letter in user_matches}
    correct_dict = {int(num): letter for num, letter in correct_matches}
    
    correct_count = 0
    wrong_count = 0
    
    for num in user_dict:
        if user_dict[num] == correct_dict.get(num):
            correct_count += 1
        else:
            wrong_count += 1
    
    return correct_count, wrong_count


def format_datetime(dt) -> str:
    if dt is None:
        return ""
    return dt.strftime("%d.%m.%Y %H:%M")


def is_valid_phone(phone: str) -> bool:
    if not phone:
        return False
    s = phone.strip().replace(' ', '').replace('-', '')
    digits = ''.join(ch for ch in s if ch.isdigit())
    return 9 <= len(digits) <= 15


def is_valid_name(name: str) -> bool:
    parts = name.strip().split()
    return len(parts) >= 2 and all(len(p) >= 2 for p in parts)


def is_valid_class(class_name: str) -> bool:
    pattern = r'^([1-9]|1[01])[-]?[A-Za-z]?$'
    return bool(re.match(pattern, class_name))


async def generate_excel_report(code: str, results: List[dict]):
    """Generates an Excel report (in-memory) for given test `results`.
    `results` is expected to be a list of dicts with keys:
    'full_name', 'phone', 'class_name', 'correct', 'wrong', 'score', 'finished_at'
    Returns: io.BytesIO object positioned at start ready to read.
    """
    sorted_results = sorted(results, key=lambda r: (-r.get('score', 0), r.get('finished_at')))

    wb = Workbook()
    ws = wb.active
    ws.title = f"Test_{code}"

    headers = ["Rank", "Ism Familiya", "Telefon", "Sinf", "To'g'ri", "Noto'g'ri", "Ball", "Sana"]
    ws.append(headers)

    for idx, row in enumerate(sorted_results, start=1):
        ws.append([
            idx,
            row.get('full_name', ''),
            row.get('phone', ''),
            row.get('class_name', ''),
            row.get('correct', 0),
            row.get('wrong', 0),
            row.get('score', 0),
            row.get('finished_at').strftime('%d.%m.%Y %H:%M') if row.get('finished_at') else ''
        ])

    for i, column_cells in enumerate(ws.columns, start=1):
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(i)].width = min(max(length + 2, 10), 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf